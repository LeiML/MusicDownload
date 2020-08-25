# __*__coding: utf-8__*__
# @Time     :2020/8/23 22:47
# @Author   :LeiLei
# @File     :MainApp.py
# @Software :PyCharm
import sys
import threading
import datetime
import requests
from lxml import etree
from openpyxl import Workbook
from PyQt5.QtWidgets import QWidget, QApplication, QMessageBox, QHeaderView, QTableView, QPushButton, QHBoxLayout, QTableWidgetItem, QFileDialog
from Frame.music import Ui_Form
from Music import *


class MusicFrame(QWidget, Ui_Form):
    def __init__(self):
        super(MusicFrame, self).__init__()
        self.setupUi(self)
        self.pushButton_2.setEnabled(False)
        self.pushButton_3.setEnabled(False)
        self.tableWidget.setColumnCount(len(TABLE_HEAD))
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.setHorizontalHeaderLabels(TABLE_HEAD)
        self.tableWidget.setEditTriggers(QTableView.NoEditTriggers)

    def search(self):
        text = self.lineEdit.text()
        if BASE_URL not in text:
            QMessageBox.warning(self, '提示', '请输入正确的URL(直接复制网页链接)', QMessageBox.Ok)
            self.lineEdit.setText('')
            return
        self.pushButton_2.setEnabled(True)
        self.pushButton_3.setEnabled(True)
        url = text.replace('/#', '')
        song_id, song_name = request_url(url)
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(len(song_id))
        for song, name in zip(song_id, song_name):
            self.tableWidget.setRowHeight(song_name.index(name), 100)
            w = QWidget()
            bnt = QPushButton('下载')
            bnt.clicked.connect(self.download)
            bnt.setStyleSheet('''
                background-color: #1010d1;
                color: white;
                font: 16pt \"宋体\";
                height: 80px;
                cursor: pointer;
            ''')
            h = QHBoxLayout()
            h.addWidget(bnt)
            w.setLayout(h)
            self.tableWidget.setCellWidget(song_name.index(name), 2, w)
            item = QTableWidgetItem(song)
            self.tableWidget.setItem(song_name.index(name), 0, item)
            item = QTableWidgetItem(name)
            self.tableWidget.setItem(song_name.index(name), 1, item)

    def download(self):
        bnt = self.sender()
        row = self.tableWidget.indexAt(bnt.parent().pos()).row()
        ids = self.tableWidget.item(row, 0).text()
        name = self.tableWidget.item(row, 1).text()
        file = QFileDialog.getSaveFileName(self, 'mp3 file', './{}.mp3'.format(name), 'mp3(*.mp3)')
        if file[0] == '':
            QMessageBox.warning(self, '提示', '文件路径无效', QMessageBox.Ok)
            return
        res = requests.get(BASE_SONG + ids + '.mp3')
        with open(file[0], 'wb')as f:
            f.write(res.content)
        QMessageBox.information(self, '提示', '{}下载成功'.format(name), QMessageBox.Ok)

    def export(self):
        url = self.lineEdit.text().replace('/#', '')
        song_id, song_name = request_url(url)
        path = QFileDialog.getSaveFileName(self, 'Excel File', './{}.xlsx'.format(get_time()), 'Excel File(*.xlsx)')
        if path[0] == '':
            QMessageBox.warning(self, '提示', '文件路径无效', QMessageBox.Ok)
            return
        book = Workbook()
        sheet = book.active
        for j in range(2):
            sheet.cell(1, j + 1, TABLE_HEAD[j])
        for i in range(len(song_id)):
            sheet.cell(i + 2, 1, song_id[i])
            sheet.cell(i + 2, 2, BASE_SONG + song_name[i])
        book.save(path[0])
        book.close()

    def download_all(self):
        if BASE_URL not in self.lineEdit.text():
            QMessageBox.warning(self, '提示', '请输入正确的URL(直接复制网页链接)', QMessageBox.Ok)
            self.lineEdit.setText('')
            return
        url = self.lineEdit.text().replace('/#', '')
        song_id, song_name = request_url(url)
        directory = QFileDialog.getExistingDirectory(self, '文件夹', '')
        if directory == '':
            QMessageBox.warning(self, '提示', '文件路径无效', QMessageBox.Ok)
            return
        thread = threading.Thread(target=download, args=(song_id, song_name, directory))
        thread.start()


def request_url(url):
    response = requests.get(url)
    html = etree.HTML(response.text)
    song_id = html.xpath('//ul[@class="f-hide"]/li/a/@href')
    song_name = html.xpath('//ul[@class="f-hide"]/li/a/text()')
    song_id = [song.replace('/song?id=', '')for song in song_id]
    return song_id, song_name


def download(ids, names, directory):
    for song, name in zip(ids, names):
        res = requests.get(BASE_SONG + song + '.mp3')
        with open(directory + '/' + name + '.mp3', 'wb')as f:
            f.write(res.content)


def get_time():
    return datetime.datetime.strftime(datetime.datetime.now(), '%Y%m%d-%H-%M-%S')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    frame = MusicFrame()
    frame.show()
    sys.exit(app.exec_())
